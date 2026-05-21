from pathlib import Path
import re
import sys

import pandas as pd
from openpyxl import Workbook


BASE_DIR = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(BASE_DIR))

from resources.halliday_intersent_dic import HALLIDAY_CONJUNCTIONS


INPUT_DIR = BASE_DIR / "outputs" / "v2_intersentential_full"
CASES_FILE = INPUT_DIR / "v2_intersentential_full_cases.csv"
TEXT_COUNTS_FILE = INPUT_DIR / "v2_intersentential_full_text_counts.csv"
DATASETS = {
    "EFCAMDAT": {
        "file": BASE_DIR / "data_filtered" / "efcamdat_v2_filtered.csv",
        "id_col": "writing_id",
        "text_col": "text_clean_preserveCase",
        "group_col": "cefr",
    },
    "COREFL": {
        "file": BASE_DIR / "data_filtered" / "corefl_v2_filtered_50.csv",
        "id_col": "text_id",
        "text_col": "text_clean_preserveCase",
        "group_col": "cefr",
    },
    "GiG": {
        "file": BASE_DIR / "data_filtered" / "gig_v2_filtered.csv",
        "id_col": "text_id",
        "text_col": "text_clean_preserveCase",
        "group_col": "year_group",
    },
}

OUT_DIR = BASE_DIR / "outputs" / "intersentential_outputs"
ADVANCED_DIR = OUT_DIR / "advanced_connector_indices"

DIAGNOSTICS_FILE = OUT_DIR / "01_diagnostics.xlsx"
MACRO_FILE = OUT_DIR / "02_macro_indices.xlsx"
SUBCATEGORY_FILE = OUT_DIR / "03_subcategory_indices.xlsx"
CASES_EVIDENCE_FILE = OUT_DIR / "04_cases_evidence.xlsx"
CONNECTOR_FILES = {
    "Extension": ADVANCED_DIR / "connector_indices_extension.xlsx",
    "Elaboration": ADVANCED_DIR / "connector_indices_elaboration.xlsx",
    "Enhancement": ADVANCED_DIR / "connector_indices_enhancement.xlsx",
}

MACROS = ["Extension", "Elaboration", "Enhancement"]
KEY_COLUMNS = ["corpus", "text_id"]
BASE_TEXT_COLUMNS = [
    "corpus",
    "text_id",
    "filename",
    "group",
    "word_count_text",
    "sentence_count_text",
    "eligible_sentence_starts",
]
CASE_EVIDENCE_COLUMNS = [
    "corpus",
    "text_id",
    "group",
    "sentence_index",
    "previous_sentence",
    "current_sentence",
    "detected_item",
    "macro_category",
    "path_2",
    "path_3",
    "path_4",
    "path_5",
    "sentence_start_cleaned",
    "intersentential_marker_type",
    "intersentential_confidence",
    "is_intersentential_supported",
    "intersentential_notes",
    "word_count_text",
    "sentence_count_text",
    "source_file",
    "algorithm_notes",
]


def slugify(text):
    text = str(text).lower().strip()
    text = text.replace("causal-conditional", "causal_conditional")
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text


def is_blank(value):
    return pd.isna(value) or str(value).strip() == ""


def clean_path_part(value):
    if is_blank(value):
        return ""
    return str(value).strip()


def split_sentences(text):
    if not isinstance(text, str):
        return []
    text = re.sub(r"\s+", " ", text.strip())
    if not text:
        return []
    parts = re.split(r"(?<=[.!?])\s+(?=[\"'“”‘’A-Z0-9])", text)
    return [part.strip() for part in parts if part.strip()]


def filename_or_text_id(df, id_col):
    filename = pd.Series(pd.NA, index=df.index, dtype="object")
    for col in ["filename", "file_name", "file_stem", "source_file", "source_path", "file_id", "file_id_norm"]:
        if col in df.columns:
            filename = filename.fillna(df[col])
    return filename.fillna(df[id_col])


def load_inputs():
    cases = pd.read_csv(CASES_FILE, low_memory=False)
    text_counts = pd.read_csv(TEXT_COUNTS_FILE, low_memory=False)

    cases["corpus"] = cases["corpus"].astype(str)
    cases["text_id"] = cases["text_id"].astype(str)
    text_counts["corpus"] = text_counts["corpus"].astype(str)
    text_counts["text_id"] = text_counts["text_id"].astype(str)
    return cases, text_counts


def build_text_metadata(corpus, spec, text_counts):
    df = pd.read_csv(spec["file"], low_memory=False)
    df[spec["id_col"]] = df[spec["id_col"]].astype(str)

    text_count_subset = text_counts[text_counts["corpus"] == corpus][
        ["text_id", "sentence_count_text"]
    ].copy()
    text_count_subset["text_id"] = text_count_subset["text_id"].astype(str)

    texts = pd.DataFrame({
        "corpus": corpus,
        "text_id": df[spec["id_col"]],
        "filename": filename_or_text_id(df, spec["id_col"]),
        "group": df[spec["group_col"]],
        "word_count_text": df["wc_preserve"],
        "sentence_count_text": df[spec["text_col"]].apply(split_sentences).apply(len),
    })
    texts = texts.drop(columns=["sentence_count_text"]).merge(
        text_count_subset,
        on="text_id",
        how="left",
    )
    if texts["sentence_count_text"].isna().any():
        fallback_counts = df[spec["text_col"]].apply(split_sentences).apply(len)
        texts["sentence_count_text"] = texts["sentence_count_text"].fillna(fallback_counts)

    texts["eligible_sentence_starts"] = (
        texts["sentence_count_text"].fillna(0).astype(int) - 1
    ).clip(lower=0)
    texts["eligible_for_intersentential"] = texts["sentence_count_text"] >= 2
    texts["diagnostic_note"] = texts["eligible_for_intersentential"].map({
        True: "Eligible: at least two sentences detected",
        False: "Not eligible: fewer than two sentences",
    })
    return texts


def build_all_text_metadata(text_counts):
    return pd.concat(
        [
            build_text_metadata(corpus, spec, text_counts)
            for corpus, spec in DATASETS.items()
        ],
        ignore_index=True,
    )


def add_rates(df, base_names):
    new_columns = {}
    for base in base_names:
        raw_col = f"{base}_raw"
        per_1000_col = f"{base}_per_1000"
        per_start_col = f"{base}_per_eligible_sentence_start"
        if raw_col in df.columns:
            raw = df[raw_col].fillna(0).astype(int)
        else:
            raw = pd.Series(0, index=df.index, dtype="int64")
        new_columns[raw_col] = raw
        new_columns[per_1000_col] = (
            raw / df["word_count_text"] * 1000
        ).where(df["word_count_text"] > 0, 0)
        new_columns[per_start_col] = (
            raw / df["eligible_sentence_starts"]
        ).where(df["eligible_sentence_starts"] > 0, 0)

    replace_cols = [col for col in new_columns if col in df.columns]
    if replace_cols:
        df = df.drop(columns=replace_cols)
    return pd.concat([df, pd.DataFrame(new_columns, index=df.index)], axis=1)


def macro_base(macro, supported=False):
    prefix = "supported_" if supported else ""
    return f"{prefix}intersentential_{slugify(macro)}"


def macro_total_base(supported=False):
    prefix = "supported_" if supported else ""
    return f"{prefix}intersentential_total"


def path_base(macro, path_parts, connector=None, supported=False):
    prefix = "supported_" if supported else ""
    parts = [macro] + [part for part in path_parts if part]
    if connector is not None:
        parts.append(connector)
    return prefix + "intersentential_" + "_".join(slugify(part) for part in parts)


def iter_dictionary_paths():
    def recurse(node, macro, path_parts):
        if isinstance(node, list):
            yield macro, path_parts, node
        elif isinstance(node, dict):
            for key, value in node.items():
                yield from recurse(value, macro, path_parts + [key])

    for macro in MACROS:
        yield from recurse(HALLIDAY_CONJUNCTIONS[macro], macro, [])


def dictionary_subcategory_bases():
    bases = []
    for macro, path_parts, _connectors in iter_dictionary_paths():
        bases.append(path_base(macro, path_parts))
    return list(dict.fromkeys(bases))


def dictionary_connector_bases_by_macro():
    bases_by_macro = {macro: [] for macro in MACROS}
    for macro, path_parts, connectors in iter_dictionary_paths():
        for connector in connectors:
            bases_by_macro[macro].append(path_base(macro, path_parts, connector))
    return {macro: list(dict.fromkeys(bases)) for macro, bases in bases_by_macro.items()}


def ordered_observed_bases(observed_bases, dictionary_bases):
    observed_bases = list(dict.fromkeys(observed_bases))
    return (
        [base for base in dictionary_bases if base in observed_bases]
        + [base for base in observed_bases if base not in dictionary_bases]
    )


def add_case_bases(cases):
    cases = cases.copy()
    path_cols = ["path_2", "path_3", "path_4", "path_5"]
    cases["macro_clean"] = cases["macro_category"].apply(clean_path_part)
    for col in path_cols:
        cases[f"{col}_clean"] = cases[col].apply(clean_path_part)
    cases["detected_item_clean"] = cases["detected_item"].apply(clean_path_part)

    cases["subcategory_base"] = cases.apply(
        lambda row: path_base(
            row["macro_clean"],
            [row[f"{col}_clean"] for col in path_cols],
        ),
        axis=1,
    )
    cases["connector_base"] = cases.apply(
        lambda row: path_base(
            row["macro_clean"],
            [row[f"{col}_clean"] for col in path_cols],
            row["detected_item_clean"],
        ),
        axis=1,
    )
    return cases


def wide_counts(cases, variable_col, expected_bases, supported=False):
    expected_cols = [f"{base}_raw" for base in expected_bases]
    if supported:
        cases = cases[cases["is_intersentential_supported"] == 1].copy()
    if cases.empty:
        return pd.DataFrame(columns=KEY_COLUMNS + expected_cols)

    counts = (
        cases.groupby(KEY_COLUMNS + [variable_col])
        .size()
        .reset_index(name="raw_count")
    )
    counts["column"] = counts[variable_col] + "_raw"
    wide = (
        counts.pivot_table(
            index=KEY_COLUMNS,
            columns="column",
            values="raw_count",
            aggfunc="sum",
            fill_value=0,
        )
        .reset_index()
    )
    wide = (
        wide.set_index(KEY_COLUMNS)
        .reindex(columns=expected_cols, fill_value=0)
        .reset_index()
    )
    return wide


def merge_count_wide(base_df, wide):
    out = base_df.merge(wide, on=KEY_COLUMNS, how="left")
    raw_cols = [col for col in out.columns if col.endswith("_raw")]
    out[raw_cols] = out[raw_cols].fillna(0).astype(int)
    return out


def build_diagnostics(all_texts, cases):
    broad_counts = (
        cases.groupby(KEY_COLUMNS)
        .size()
        .reset_index(name="intersentential_total_raw")
    )
    supported_counts = (
        cases[cases["is_intersentential_supported"] == 1]
        .groupby(KEY_COLUMNS)
        .size()
        .reset_index(name="supported_intersentential_total_raw")
    )
    diagnostics = all_texts.merge(broad_counts, on=KEY_COLUMNS, how="left")
    diagnostics = diagnostics.merge(supported_counts, on=KEY_COLUMNS, how="left")
    for col in ["intersentential_total_raw", "supported_intersentential_total_raw"]:
        diagnostics[col] = diagnostics[col].fillna(0).astype(int)
    return diagnostics[[
        "corpus",
        "text_id",
        "filename",
        "group",
        "word_count_text",
        "sentence_count_text",
        "eligible_sentence_starts",
        "eligible_for_intersentential",
        "diagnostic_note",
        "intersentential_total_raw",
        "supported_intersentential_total_raw",
    ]]


def build_macro_indices(eligible_texts, cases):
    out = eligible_texts[BASE_TEXT_COLUMNS].copy()

    broad_counts = (
        cases.groupby(KEY_COLUMNS + ["macro_clean"])
        .size()
        .reset_index(name="raw_count")
    )
    supported_counts = (
        cases[cases["is_intersentential_supported"] == 1]
        .groupby(KEY_COLUMNS + ["macro_clean"])
        .size()
        .reset_index(name="raw_count")
    )

    for macro in MACROS:
        base = macro_base(macro)
        values = broad_counts[broad_counts["macro_clean"] == macro][KEY_COLUMNS + ["raw_count"]]
        values = values.rename(columns={"raw_count": f"{base}_raw"})
        out = out.merge(values, on=KEY_COLUMNS, how="left")

        supported_base = macro_base(macro, supported=True)
        supported_values = supported_counts[
            supported_counts["macro_clean"] == macro
        ][KEY_COLUMNS + ["raw_count"]]
        supported_values = supported_values.rename(
            columns={"raw_count": f"{supported_base}_raw"}
        )
        out = out.merge(supported_values, on=KEY_COLUMNS, how="left")

    broad_bases = [macro_base(macro) for macro in MACROS]
    supported_bases = [macro_base(macro, supported=True) for macro in MACROS]
    raw_cols = [f"{base}_raw" for base in broad_bases + supported_bases]
    out[raw_cols] = out[raw_cols].fillna(0).astype(int)

    out[f"{macro_total_base()}_raw"] = out[[f"{base}_raw" for base in broad_bases]].sum(axis=1)
    out[f"{macro_total_base(True)}_raw"] = out[
        [f"{base}_raw" for base in supported_bases]
    ].sum(axis=1)

    ordered_bases = [macro_total_base()] + broad_bases + [macro_total_base(True)] + supported_bases
    out = add_rates(out, ordered_bases)

    ordered_cols = list(BASE_TEXT_COLUMNS)
    for base in [macro_total_base()] + broad_bases:
        ordered_cols += [
            f"{base}_raw",
            f"{base}_per_1000",
            f"{base}_per_eligible_sentence_start",
        ]
    for base in [macro_total_base(True)] + supported_bases:
        ordered_cols += [
            f"{base}_raw",
            f"{base}_per_1000",
            f"{base}_per_eligible_sentence_start",
        ]
    return out[ordered_cols]


def build_subcategory_indices(eligible_texts, cases):
    base_df = eligible_texts[BASE_TEXT_COLUMNS].copy()
    dictionary_bases = dictionary_subcategory_bases()
    observed_bases = list(dict.fromkeys(cases["subcategory_base"].dropna()))
    broad_bases = ordered_observed_bases(observed_bases, dictionary_bases)
    supported_bases = [f"supported_{base}" for base in broad_bases]

    broad_wide = wide_counts(cases, "subcategory_base", broad_bases)
    supported_cases = cases.copy()
    supported_cases["supported_subcategory_base"] = "supported_" + supported_cases["subcategory_base"]
    supported_wide = wide_counts(
        supported_cases,
        "supported_subcategory_base",
        supported_bases,
        supported=True,
    )

    out = merge_count_wide(base_df, broad_wide)
    out = merge_count_wide(out, supported_wide)
    out = add_rates(out, broad_bases + supported_bases)

    ordered_cols = list(BASE_TEXT_COLUMNS)
    for base in broad_bases + supported_bases:
        ordered_cols += [
            f"{base}_raw",
            f"{base}_per_1000",
            f"{base}_per_eligible_sentence_start",
        ]
    return out[ordered_cols]


def build_cases_evidence(cases):
    evidence = cases.copy()
    for col in CASE_EVIDENCE_COLUMNS:
        if col not in evidence.columns:
            evidence[col] = pd.NA
    return evidence[CASE_EVIDENCE_COLUMNS]


def build_connector_indices(eligible_texts, cases, macro, dictionary_bases_by_macro):
    base_df = eligible_texts[BASE_TEXT_COLUMNS].copy()
    macro_cases = cases[cases["macro_clean"] == macro].copy()
    dictionary_bases = dictionary_bases_by_macro[macro]
    observed_bases = list(dict.fromkeys(macro_cases["connector_base"].dropna()))
    broad_bases = ordered_observed_bases(observed_bases, dictionary_bases)
    supported_bases = [f"supported_{base}" for base in broad_bases]

    broad_wide = wide_counts(macro_cases, "connector_base", broad_bases)
    supported_cases = macro_cases.copy()
    supported_cases["supported_connector_base"] = "supported_" + supported_cases["connector_base"]
    supported_wide = wide_counts(
        supported_cases,
        "supported_connector_base",
        supported_bases,
        supported=True,
    )

    out = merge_count_wide(base_df, broad_wide)
    out = merge_count_wide(out, supported_wide)
    out = add_rates(out, broad_bases + supported_bases)

    ordered_cols = list(BASE_TEXT_COLUMNS)
    for base in broad_bases + supported_bases:
        ordered_cols += [
            f"{base}_raw",
            f"{base}_per_1000",
            f"{base}_per_eligible_sentence_start",
        ]
    return out[ordered_cols]


def write_excel(df, path, sheet_name):
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title=sheet_name)
    ws.freeze_panes = "A2"
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append(row)
    wb.save(path)


def safe_rate(raw, denominator, multiplier=1):
    if pd.isna(denominator) or denominator <= 0:
        return 0
    return raw / denominator * multiplier


def count_map(cases, variable_col, supported=False):
    if supported:
        cases = cases[cases["is_intersentential_supported"] == 1].copy()
    if cases.empty:
        return {}
    counts = (
        cases.groupby(KEY_COLUMNS + [variable_col])
        .size()
        .reset_index(name="raw_count")
    )
    return {
        (row.corpus, row.text_id, getattr(row, variable_col)): row.raw_count
        for row in counts.itertuples(index=False)
    }


def stream_count_index(eligible_texts, cases, variable_col, bases, path, sheet_name):
    path.parent.mkdir(parents=True, exist_ok=True)
    broad_counts = count_map(cases, variable_col)
    supported_counts = count_map(cases, variable_col, supported=True)
    supported_bases = [f"supported_{base}" for base in bases]

    headers = list(BASE_TEXT_COLUMNS)
    for base in bases + supported_bases:
        headers.extend([
            f"{base}_raw",
            f"{base}_per_1000",
            f"{base}_per_eligible_sentence_start",
        ])

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title=sheet_name)
    ws.freeze_panes = "A2"
    ws.append(headers)

    for row in eligible_texts[BASE_TEXT_COLUMNS].itertuples(index=False):
        row_values = list(row)
        corpus = row.corpus
        text_id = row.text_id
        word_count = row.word_count_text
        eligible_starts = row.eligible_sentence_starts

        for base in bases:
            raw = broad_counts.get((corpus, text_id, base), 0)
            if raw:
                row_values.extend([
                    raw,
                    safe_rate(raw, word_count, 1000),
                    safe_rate(raw, eligible_starts),
                ])
            else:
                row_values.extend([None, None, None])

        for base in bases:
            raw = supported_counts.get((corpus, text_id, base), 0)
            if raw:
                row_values.extend([
                    raw,
                    safe_rate(raw, word_count, 1000),
                    safe_rate(raw, eligible_starts),
                ])
            else:
                row_values.extend([None, None, None])

        ws.append(row_values)

    wb.save(path)
    return len(eligible_texts), len(headers)


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    ADVANCED_DIR.mkdir(parents=True, exist_ok=True)

    cases, text_counts = load_inputs()
    cases = add_case_bases(cases)

    all_texts = build_all_text_metadata(text_counts)
    eligible_texts = all_texts[all_texts["eligible_for_intersentential"]].copy()

    diagnostics = build_diagnostics(all_texts, cases)
    macro_indices = build_macro_indices(eligible_texts, cases)
    cases_evidence = build_cases_evidence(cases)

    connector_bases_by_macro = dictionary_connector_bases_by_macro()
    dictionary_subcategory = dictionary_subcategory_bases()
    observed_subcategory = list(dict.fromkeys(cases["subcategory_base"].dropna()))
    subcategory_bases = ordered_observed_bases(observed_subcategory, dictionary_subcategory)

    write_excel(diagnostics, DIAGNOSTICS_FILE, "Diagnostics")
    write_excel(macro_indices, MACRO_FILE, "Macro indices")
    subcategory_shape = stream_count_index(
        eligible_texts,
        cases,
        "subcategory_base",
        subcategory_bases,
        SUBCATEGORY_FILE,
        "Subcategory indices",
    )
    write_excel(cases_evidence, CASES_EVIDENCE_FILE, "Cases evidence")

    connector_shapes = {}
    for macro in MACROS:
        macro_cases = cases[cases["macro_clean"] == macro].copy()
        observed_connector = list(dict.fromkeys(macro_cases["connector_base"].dropna()))
        connector_bases = ordered_observed_bases(
            observed_connector,
            connector_bases_by_macro[macro],
        )
        connector_shapes[macro] = stream_count_index(
            eligible_texts,
            macro_cases,
            "connector_base",
            connector_bases,
            CONNECTOR_FILES[macro],
            f"{macro} connectors",
        )

    output_shapes = {
        DIAGNOSTICS_FILE: diagnostics.shape,
        MACRO_FILE: macro_indices.shape,
        SUBCATEGORY_FILE: subcategory_shape,
        CASES_EVIDENCE_FILE: cases_evidence.shape,
    }
    output_shapes.update({
        CONNECTOR_FILES[macro]: connector_shapes[macro]
        for macro in MACROS
    })

    print("Built inter-sentential outcome package")
    print("Total texts:", len(all_texts))
    print("Eligible texts:", len(eligible_texts))
    print("Case rows:", len(cases))
    print("\nOutput files:")
    for path, (n_rows, n_cols) in output_shapes.items():
        print(f"- {path}: {n_rows} rows x {n_cols} columns")


if __name__ == "__main__":
    main()
