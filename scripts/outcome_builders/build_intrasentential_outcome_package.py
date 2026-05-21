from pathlib import Path
import re
import sys

import pandas as pd
from openpyxl import Workbook


BASE_DIR = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(BASE_DIR))

from resources.halliday_intrasent_dic import INTRA_CLAUSE_CONJUNCTIONS


INPUT_DIR = BASE_DIR / "outputs" / "v2_intrasentential_full"
CASES_FILE = INPUT_DIR / "v2_intrasentential_full_cases.csv"
TEXT_INDICES_FILE = INPUT_DIR / "v2_intrasentential_full_text_indices.csv"
SUPPORTED_TEXT_INDICES_FILE = INPUT_DIR / "v2_intrasentential_full_supported_text_indices.csv"
DATASETS = {
    "EFCAMDAT": {
        "file": BASE_DIR / "data_filtered" / "efcamdat_v2_filtered.csv",
        "id_col": "writing_id",
        "text_col": "text_clean_preserveCase",
        "group_col": "cefr",
    },
    "COREFL": {
        "file": BASE_DIR / "data_filtered" / "corefl_v2_filtered.csv",
        "id_col": "text_id",
        "text_col": "text_clean_preserveCase",
        "group_col": "cefr",
    },
    "GiG": {
        "file": BASE_DIR / "data_filtered" / "gig_v2_filtered.csv",
        "id_col": "text_id",
        "text_col": "text_clean_preserveCase",
        "group_col": "year_group",
    },
}

OUT_DIR = BASE_DIR / "outputs" / "intrasentential_outputs"
ADVANCED_DIR = OUT_DIR / "advanced_connector_indices"

DIAGNOSTICS_FILE = OUT_DIR / "01_diagnostics.xlsx"
MACRO_TAXIS_FILE = OUT_DIR / "02_macro_taxis_indices.xlsx"
SUBCATEGORY_FILE = OUT_DIR / "03_subcategory_indices.xlsx"
CASES_EVIDENCE_FILE = OUT_DIR / "04_cases_evidence.csv"
CONNECTOR_FILES = {
    "Extension": ADVANCED_DIR / "connector_indices_extension.xlsx",
    "Elaboration": ADVANCED_DIR / "connector_indices_elaboration.xlsx",
    "Enhancement": ADVANCED_DIR / "connector_indices_enhancement.xlsx",
}

MACROS = ["Extension", "Elaboration", "Enhancement"]
TAXES = ["paratactic", "hypotactic"]
KEY_COLUMNS = ["corpus", "text_id"]
BASE_TEXT_COLUMNS = [
    "corpus",
    "text_id",
    "filename",
    "group",
    "word_count_text",
    "sentence_count_text",
]
CASE_EVIDENCE_COLUMNS = [
    "corpus",
    "text_id",
    "group",
    "sentence_index",
    "detected_item",
    "macro_category",
    "path_2",
    "path_3",
    "path_4",
    "path_5",
    "connector_start_char",
    "connector_end_char",
    "sentence",
    "taxis",
    "is_problematic_multifunctional",
    "priming_decision",
    "priming_confidence",
    "priming_notes",
    "is_priming_supported",
    "word_count_text",
    "sentence_count_text",
    "source_file",
    "algorithm_notes",
]


def slugify(text):
    text = str(text).lower().strip()
    text = text.replace("causal-conditional", "causal_conditional")
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text


def is_blank(value):
    return pd.isna(value) or str(value).strip() == ""


def clean_path_part(value):
    if is_blank(value):
        return ""
    return str(value).strip()


def split_sentences(text):
    if not isinstance(text, str):
        return []
    text = re.sub(r"\s+", " ", text.strip())
    if not text:
        return []
    parts = re.split(r"(?<=[.!?])\s+(?=[\"'“”‘’A-Z0-9])", text)
    return [part.strip() for part in parts if part.strip()]


def filename_or_text_id(df, id_col):
    filename = pd.Series(pd.NA, index=df.index, dtype="object")
    for col in ["filename", "file_name", "file_stem", "source_file", "source_path", "file_id", "file_id_norm"]:
        if col in df.columns:
            filename = filename.fillna(df[col])
    return filename.fillna(df[id_col])


def load_inputs():
    cases = pd.read_csv(CASES_FILE, low_memory=False)
    text_indices = pd.read_csv(TEXT_INDICES_FILE, low_memory=False)
    supported_text_indices = pd.read_csv(SUPPORTED_TEXT_INDICES_FILE, low_memory=False)

    for df in [cases, text_indices, supported_text_indices]:
        df["corpus"] = df["corpus"].astype(str)
        df["text_id"] = df["text_id"].astype(str)
    return cases, text_indices, supported_text_indices


def build_text_metadata(corpus, spec, text_indices):
    df = pd.read_csv(spec["file"], low_memory=False)
    df[spec["id_col"]] = df[spec["id_col"]].astype(str)
    text_count_subset = text_indices[text_indices["corpus"] == corpus][
        ["text_id", "sentence_count_text"]
    ].copy()
    text_count_subset["text_id"] = text_count_subset["text_id"].astype(str)

    texts = pd.DataFrame({
        "corpus": corpus,
        "text_id": df[spec["id_col"]],
        "filename": filename_or_text_id(df, spec["id_col"]),
        "group": df[spec["group_col"]],
        "word_count_text": df["wc_preserve"],
        "sentence_count_text": df[spec["text_col"]].apply(split_sentences).apply(len),
    })
    texts = texts.drop(columns=["sentence_count_text"]).merge(
        text_count_subset,
        on="text_id",
        how="left",
    )
    if texts["sentence_count_text"].isna().any():
        fallback_counts = df[spec["text_col"]].apply(split_sentences).apply(len)
        texts["sentence_count_text"] = texts["sentence_count_text"].fillna(fallback_counts)
    texts["eligible_for_intrasentential"] = texts["word_count_text"].fillna(0) > 0
    texts["diagnostic_note"] = texts["eligible_for_intrasentential"].map({
        True: "Eligible: text contains words",
        False: "Not eligible: missing or zero word count",
    })
    return texts


def build_all_text_metadata(text_indices):
    return pd.concat(
        [
            build_text_metadata(corpus, spec, text_indices)
            for corpus, spec in DATASETS.items()
        ],
        ignore_index=True,
    )


def safe_rate(raw, word_count):
    if pd.isna(word_count) or word_count <= 0:
        return 0
    return raw / word_count * 1000


def add_rates(df, base_names):
    new_columns = {}
    for base in base_names:
        raw_col = f"{base}_raw"
        per_1000_col = f"{base}_per_1000"
        if raw_col in df.columns:
            raw = df[raw_col].fillna(0).astype(int)
        else:
            raw = pd.Series(0, index=df.index, dtype="int64")
        new_columns[raw_col] = raw
        new_columns[per_1000_col] = (
            raw / df["word_count_text"] * 1000
        ).where(df["word_count_text"] > 0, 0)

    replace_cols = [col for col in new_columns if col in df.columns]
    if replace_cols:
        df = df.drop(columns=replace_cols)
    return pd.concat([df, pd.DataFrame(new_columns, index=df.index)], axis=1)


def path_base(taxis, macro, path_parts, connector=None, supported=False):
    prefix = "supported_" if supported else ""
    taxis = clean_path_part(taxis) or "unspecified"
    parts = [taxis, macro] + [
        part for part in path_parts
        if part and part not in {"paratactic", "hypotactic"}
    ]
    if connector is not None:
        parts.append(connector)
    return prefix + "intrasentential_" + "_".join(slugify(part) for part in parts)


def iter_dictionary_paths():
    def recurse(node, macro, path_parts):
        if isinstance(node, list):
            taxis = "unspecified"
            labels = []
            for part in path_parts:
                if part in {"paratactic", "hypotactic"}:
                    taxis = part
                else:
                    labels.append(part)
            yield macro, taxis, labels[1:], node
        elif isinstance(node, dict):
            for key, value in node.items():
                yield from recurse(value, macro, path_parts + [key])

    for macro in MACROS:
        yield from recurse(INTRA_CLAUSE_CONJUNCTIONS[macro], macro, [macro])


def dictionary_subcategory_bases():
    bases = []
    for macro, taxis, path_parts, _connectors in iter_dictionary_paths():
        bases.append(path_base(taxis, macro, path_parts))
    return list(dict.fromkeys(bases))


def dictionary_connector_bases_by_macro():
    bases_by_macro = {macro: [] for macro in MACROS}
    for macro, taxis, path_parts, connectors in iter_dictionary_paths():
        for connector in connectors:
            bases_by_macro[macro].append(path_base(taxis, macro, path_parts, connector))
    return {macro: list(dict.fromkeys(bases)) for macro, bases in bases_by_macro.items()}


def ordered_observed_bases(observed_bases, dictionary_bases):
    observed_bases = list(dict.fromkeys(observed_bases))
    return (
        [base for base in dictionary_bases if base in observed_bases]
        + [base for base in observed_bases if base not in dictionary_bases]
    )


def add_case_bases(cases):
    cases = cases.copy()
    path_cols = ["path_2", "path_3", "path_4", "path_5"]
    cases["macro_clean"] = cases["macro_category"].apply(clean_path_part)
    cases["taxis_clean"] = cases["taxis"].apply(clean_path_part).replace("", "unspecified")
    for col in path_cols:
        cases[f"{col}_clean"] = cases[col].apply(clean_path_part)
    cases["detected_item_clean"] = cases["detected_item"].apply(clean_path_part)

    cases["subcategory_base"] = cases.apply(
        lambda row: path_base(
            row["taxis_clean"],
            row["macro_clean"],
            [row[f"{col}_clean"] for col in path_cols],
        ),
        axis=1,
    )
    cases["connector_base"] = cases.apply(
        lambda row: path_base(
            row["taxis_clean"],
            row["macro_clean"],
            [row[f"{col}_clean"] for col in path_cols],
            row["detected_item_clean"],
        ),
        axis=1,
    )
    return cases


def build_diagnostics(all_texts, cases):
    broad_counts = (
        cases.groupby(KEY_COLUMNS)
        .size()
        .reset_index(name="intra_total_raw")
    )
    supported_counts = (
        cases[cases["is_priming_supported"] == 1]
        .groupby(KEY_COLUMNS)
        .size()
        .reset_index(name="supported_intra_total_raw")
    )
    diagnostics = all_texts.merge(broad_counts, on=KEY_COLUMNS, how="left")
    diagnostics = diagnostics.merge(supported_counts, on=KEY_COLUMNS, how="left")
    for col in ["intra_total_raw", "supported_intra_total_raw"]:
        diagnostics[col] = diagnostics[col].fillna(0).astype(int)
    return diagnostics[[
        "corpus",
        "text_id",
        "filename",
        "group",
        "word_count_text",
        "sentence_count_text",
        "eligible_for_intrasentential",
        "diagnostic_note",
        "intra_total_raw",
        "supported_intra_total_raw",
    ]]


def macro_taxis_base(taxis=None, macro=None, supported=False):
    prefix = "supported_" if supported else ""
    if taxis and macro is None:
        return f"{prefix}intra_{taxis}_total"
    if macro:
        if taxis:
            return f"{prefix}intra_{taxis}_{slugify(macro)}"
        return f"{prefix}intra_{slugify(macro)}"
    return f"{prefix}intra_total"


def build_macro_taxis_indices(eligible_texts, cases):
    out = eligible_texts[BASE_TEXT_COLUMNS].copy()
    broad_counts = (
        cases.groupby(KEY_COLUMNS + ["taxis_clean", "macro_clean"])
        .size()
        .reset_index(name="raw_count")
    )
    supported_counts = (
        cases[cases["is_priming_supported"] == 1]
        .groupby(KEY_COLUMNS + ["taxis_clean", "macro_clean"])
        .size()
        .reset_index(name="raw_count")
    )

    for supported, counts in [(False, broad_counts), (True, supported_counts)]:
        for taxis in TAXES:
            for macro in MACROS:
                base = macro_taxis_base(taxis, macro, supported)
                values = counts[
                    (counts["taxis_clean"] == taxis)
                    & (counts["macro_clean"] == macro)
                ][KEY_COLUMNS + ["raw_count"]]
                values = values.rename(columns={"raw_count": f"{base}_raw"})
                out = out.merge(values, on=KEY_COLUMNS, how="left")

    broad_bases = [
        macro_taxis_base(),
        macro_taxis_base("paratactic"),
        macro_taxis_base("hypotactic"),
    ] + [
        macro_taxis_base(taxis, macro)
        for taxis in TAXES
        for macro in MACROS
    ]
    supported_bases = [
        macro_taxis_base(supported=True),
        macro_taxis_base("paratactic", supported=True),
        macro_taxis_base("hypotactic", supported=True),
    ] + [
        macro_taxis_base(taxis, macro, supported=True)
        for taxis in TAXES
        for macro in MACROS
    ]

    component_broad = [
        macro_taxis_base(taxis, macro)
        for taxis in TAXES
        for macro in MACROS
    ]
    component_supported = [
        macro_taxis_base(taxis, macro, supported=True)
        for taxis in TAXES
        for macro in MACROS
    ]
    raw_cols = [f"{base}_raw" for base in component_broad + component_supported]
    out[raw_cols] = out[raw_cols].fillna(0).astype(int)

    for taxis in TAXES:
        out[f"{macro_taxis_base(taxis)}_raw"] = out[
            [f"{macro_taxis_base(taxis, macro)}_raw" for macro in MACROS]
        ].sum(axis=1)
        out[f"{macro_taxis_base(taxis, supported=True)}_raw"] = out[
            [f"{macro_taxis_base(taxis, macro, supported=True)}_raw" for macro in MACROS]
        ].sum(axis=1)

    out[f"{macro_taxis_base()}_raw"] = out[
        [f"{macro_taxis_base(taxis)}_raw" for taxis in TAXES]
    ].sum(axis=1)
    out[f"{macro_taxis_base(supported=True)}_raw"] = out[
        [f"{macro_taxis_base(taxis, supported=True)}_raw" for taxis in TAXES]
    ].sum(axis=1)

    out = add_rates(out, broad_bases + supported_bases)

    ordered_cols = list(BASE_TEXT_COLUMNS)
    for base in broad_bases + supported_bases:
        ordered_cols += [f"{base}_raw", f"{base}_per_1000"]
    return out[ordered_cols]


def count_map(cases, variable_col, supported=False):
    if supported:
        cases = cases[cases["is_priming_supported"] == 1].copy()
    if cases.empty:
        return {}
    counts = (
        cases.groupby(KEY_COLUMNS + [variable_col])
        .size()
        .reset_index(name="raw_count")
    )
    return {
        (row.corpus, row.text_id, getattr(row, variable_col)): row.raw_count
        for row in counts.itertuples(index=False)
    }


def stream_count_index(eligible_texts, cases, variable_col, bases, path, sheet_name):
    path.parent.mkdir(parents=True, exist_ok=True)
    broad_counts = count_map(cases, variable_col)
    supported_counts = count_map(cases, variable_col, supported=True)
    supported_bases = [f"supported_{base}" for base in bases]

    headers = list(BASE_TEXT_COLUMNS)
    for base in bases + supported_bases:
        headers.extend([f"{base}_raw", f"{base}_per_1000"])

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title=sheet_name)
    ws.freeze_panes = "A2"
    ws.append(headers)

    for row in eligible_texts[BASE_TEXT_COLUMNS].itertuples(index=False):
        row_values = list(row)
        corpus = row.corpus
        text_id = row.text_id
        word_count = row.word_count_text

        for base in bases:
            raw = broad_counts.get((corpus, text_id, base), 0)
            if raw:
                row_values.extend([raw, safe_rate(raw, word_count)])
            else:
                row_values.extend([None, None])

        for base in bases:
            raw = supported_counts.get((corpus, text_id, base), 0)
            if raw:
                row_values.extend([raw, safe_rate(raw, word_count)])
            else:
                row_values.extend([None, None])

        ws.append(row_values)

    wb.save(path)
    return len(eligible_texts), len(headers)


def build_cases_evidence(cases):
    evidence = cases.copy()
    for col in CASE_EVIDENCE_COLUMNS:
        if col not in evidence.columns:
            evidence[col] = pd.NA
    return evidence[CASE_EVIDENCE_COLUMNS]


def write_excel(df, path, sheet_name):
    path.parent.mkdir(parents=True, exist_ok=True)
    max_data_rows = 1_048_575
    wb = Workbook(write_only=True)

    for sheet_idx, start in enumerate(range(0, len(df), max_data_rows), start=1):
        if sheet_idx == 1:
            current_sheet_name = sheet_name[:31]
        else:
            suffix = f" {sheet_idx:02d}"
            current_sheet_name = sheet_name[:31 - len(suffix)] + suffix
        ws = wb.create_sheet(title=current_sheet_name)
        ws.freeze_panes = "A2"
        ws.append(list(df.columns))
        chunk = df.iloc[start:start + max_data_rows]
        for row in chunk.itertuples(index=False, name=None):
            ws.append(row)

    if len(df) == 0:
        ws = wb.create_sheet(title=sheet_name[:31])
        ws.freeze_panes = "A2"
        ws.append(list(df.columns))

    wb.save(path)


def write_csv(df, path):
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(path, index=False)


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    ADVANCED_DIR.mkdir(parents=True, exist_ok=True)

    cases, text_indices, supported_text_indices = load_inputs()
    cases = add_case_bases(cases)
    all_texts = build_all_text_metadata(text_indices)
    eligible_texts = all_texts[all_texts["eligible_for_intrasentential"]].copy()

    diagnostics = build_diagnostics(all_texts, cases)
    macro_taxis_indices = build_macro_taxis_indices(eligible_texts, cases)
    cases_evidence = build_cases_evidence(cases)

    dictionary_subcategory = dictionary_subcategory_bases()
    observed_subcategory = list(dict.fromkeys(cases["subcategory_base"].dropna()))
    subcategory_bases = ordered_observed_bases(observed_subcategory, dictionary_subcategory)
    connector_bases_by_macro = dictionary_connector_bases_by_macro()

    write_excel(diagnostics, DIAGNOSTICS_FILE, "Diagnostics")
    write_excel(macro_taxis_indices, MACRO_TAXIS_FILE, "Macro taxis indices")
    subcategory_shape = stream_count_index(
        eligible_texts,
        cases,
        "subcategory_base",
        subcategory_bases,
        SUBCATEGORY_FILE,
        "Subcategory indices",
    )
    write_csv(cases_evidence, CASES_EVIDENCE_FILE)

    connector_shapes = {}
    for macro in MACROS:
        macro_cases = cases[cases["macro_clean"] == macro].copy()
        observed_connector = list(dict.fromkeys(macro_cases["connector_base"].dropna()))
        connector_bases = ordered_observed_bases(
            observed_connector,
            connector_bases_by_macro[macro],
        )
        connector_shapes[macro] = stream_count_index(
            eligible_texts,
            macro_cases,
            "connector_base",
            connector_bases,
            CONNECTOR_FILES[macro],
            f"{macro} connectors",
        )

    output_shapes = {
        DIAGNOSTICS_FILE: diagnostics.shape,
        MACRO_TAXIS_FILE: macro_taxis_indices.shape,
        SUBCATEGORY_FILE: subcategory_shape,
        CASES_EVIDENCE_FILE: cases_evidence.shape,
    }
    output_shapes.update({
        CONNECTOR_FILES[macro]: connector_shapes[macro]
        for macro in MACROS
    })

    print("Built intra-sentential outcome package")
    print("Total texts:", len(all_texts))
    print("Eligible texts:", len(eligible_texts))
    print("Case rows:", len(cases))
    print("\nOutput files:")
    for path, (n_rows, n_cols) in output_shapes.items():
        print(f"- {path}: {n_rows} rows x {n_cols} columns")


if __name__ == "__main__":
    main()
